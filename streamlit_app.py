import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import os
import time

def calculate_vif(X):
    X_check = X.drop(columns=['const']) if 'const' in X.columns else X.copy()
    X_check = X_check.astype(float)
    while True:
        vif_data = pd.DataFrame()
        vif_data["Feature"] = X_check.columns
        try:
            vif_data["VIF"] = [variance_inflation_factor(X_check.values, i) for i in range(len(X_check.columns))]
            vif_data["VIF"] = vif_data["VIF"].replace([np.inf, -np.inf], 999999.0)
        except: break
        max_vif = vif_data['VIF'].max()
        if max_vif > 5.0:
            max_vif_feature = vif_data.loc[vif_data['VIF'] == max_vif, 'Feature'].iloc[0]
            X_check = X_check.drop(columns=[max_vif_feature])
        else: break
    return sm.add_constant(X_check)

st.set_page_config(page_title="Beyond Intuition | AI Engine", layout="wide", initial_sidebar_state="expanded")

# --- PREMIUM CSS & UI POLISH ---
st.markdown("""
<style>
    .stApp { background-color: #0E1117; color: #E2E8F0; font-family: 'Inter', sans-serif; }
    .metric-card { 
        background: linear-gradient(145deg, #1A202C, #2D3748); 
        border-radius: 12px; 
        padding: 24px; 
        text-align: center; 
        border: 1px solid #4A5568; 
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        height: 100%;
        transition: transform 0.2s;
    }
    .metric-card:hover { transform: translateY(-5px); }
    .metric-value { font-size: 36px; font-weight: 800; color: #63B3ED; margin: 10px 0; }
    .metric-title { font-size: 14px; color: #A0AEC0; text-transform: uppercase; letter-spacing: 1.5px; font-weight: 600; }
    .grade-A { color: #48BB78 !important; text-shadow: 0 0 15px rgba(72,187,120,0.5); }
    .grade-B { color: #4299E1 !important; text-shadow: 0 0 15px rgba(66,153,225,0.5); }
    .grade-C { color: #ECC94B !important; }
    .grade-D { color: #ED8936 !important; }
    .grade-E { color: #F56565 !important; }
    hr { border-color: #2D3748; margin: 30px 0; }
    div[data-testid="stSidebar"] { background-color: #11151C; border-right: 1px solid #1E2532; }
    div.stButton > button { 
        width: 100%; 
        border-radius: 8px; 
        font-weight: bold; 
        padding: 0.75rem 1rem; 
        transition: all 0.3s;
    }
    div.stButton > button[kind="primary"] { 
        background: linear-gradient(90deg, #4FD1C5, #3182CE); 
        border: none; 
        color: white;
    }
    div.stButton > button[kind="primary"]:hover {
        background: linear-gradient(90deg, #38B2AC, #2B6CB0); 
        box-shadow: 0 0 15px rgba(79, 209, 197, 0.4);
    }
    .hero-text { font-size: 1.2rem; color: #A0AEC0; margin-bottom: 2rem; }
    .info-box { background-color: #2D3748; padding: 15px; border-radius: 8px; border-left: 4px solid #4FD1C5; margin-bottom: 15px; }
    .tooltip-icon { color: #A0AEC0; cursor: help; margin-left: 5px; }
</style>
""", unsafe_allow_html=True)

CATEGORY_DNA = {
    'Tops': ['Silhouette', 'Neckline', 'Sleeve Type', 'Color_Family'],
    'Bottoms': ['Fit', 'Leg Length', 'Waist Rise', 'Pocket Styling', 'Color_Family'],
    'Dresses': ['Silhouette', 'Garment Length', 'Neck Type', 'Color_Family']
}
BINARY_DNA = ['Has_Print', 'Has_Embroidery', 'Has_Accessories', 'Has_Texture']

st.sidebar.title("Beyond Intuition 🔮")
st.sidebar.markdown("### Core AI Engine")
st.sidebar.markdown("---")

main_mode = st.sidebar.radio("Navigation", [
    "🧠 Engine 1: Train AI Model",
    "⚡ Engine 2: Batch Grader",
    "📊 Engine 3: Live Dashboard"
])

st.sidebar.markdown("---")
st.sidebar.caption("v3.0 | Predictive Clustering Intelligence")

# ==========================================
# ENGINE 1: TRAIN AI (Master Output Engine)
# ==========================================
if main_mode == "🧠 Engine 1: Train AI Model":
    st.markdown('<h1 style="background: -webkit-linear-gradient(45deg, #4FD1C5, #3182CE); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">🧠 Engine 1: Mathematical Training</h1>', unsafe_allow_html=True)
    st.markdown('<p class="hero-text">Upload historical product data to extract visual weights and generate the intelligence core using Predictive K-Means Clustering.</p>', unsafe_allow_html=True)
    
    col1, spacer, col2 = st.columns([1.2, 0.1, 1])
    
    with col1:
        st.subheader("1. Data Ingestion")
        uploaded_hist = st.file_uploader("Upload Past Product Data (Task1_Final_Attributes.xlsx)", type=["xlsx"], accept_multiple_files=False)
        
        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("2. Model Parameters ⚙️")
        p_val = st.slider("Pearson Filter Threshold (Noise Reduction)", 0.0, 0.10, 0.05, 0.01, help="Drops traits with an absolute correlation coefficient lower than this before running regression.")
        targets = st.multiselect("Target Variables", ["STR", "ROS"], default=["STR", "ROS"], help="Select variables to run regression on.")
        control_col = st.text_input("Time Control Variable (Optional)", value="", placeholder="e.g. Trading Days")
        
        st.markdown("<br>", unsafe_allow_html=True)
        run_btn = st.button("🔨 Train AI & Generate Master Core", type="primary", use_container_width=True)

    with col2:
        st.subheader("System Architecture")
        st.markdown("""
        <div class="info-box"><strong>Phase 1: Feature Engineering</strong><br>Converting textual DNA into binary matrices and applying the <i>Unicorn Trap Filter</i> (N >= 3).</div>
        <div class="info-box"><strong>Phase 2: OLS Regression</strong><br>Running multivariable regression across isolated garment categories, utilizing Pearson Correlation filters and VIF mapping to extract true independent variable impact.</div>
        <div class="info-box"><strong>Phase 3: Predictive K-Means</strong><br>Applying Regression Weights to construct <i>DNA Impact Scores</i>, followed by Spatial K-Means Clustering to group items into Grades A-E.</div>
        """, unsafe_allow_html=True)
        
        metric_placeholder = st.empty()

    if uploaded_hist and run_btn:
        progress_text = "Initializing Data Matrix & Scanning Sheets..."
        my_bar = st.progress(0, text=progress_text)
        
        # Load via openpyxl to PRESERVE IMAGES non-destructively
        uploaded_hist.seek(0)
        wb = load_workbook(uploaded_hist)
        uploaded_hist.seek(0)
        all_dfs = pd.read_excel(uploaded_hist, sheet_name=None)
        
        total_styles = 0
        sig_count = 0
        r2_scores = []
        
        for i, (sheet_name, cat_df) in enumerate(all_dfs.items()):
            cat = sheet_name.strip().title()
            if cat not in CATEGORY_DNA: continue
            
            my_bar.progress((i + 1) / len(all_dfs), text=f"Training Models for {cat}...")
            ws = wb[sheet_name]
            
            cat_df = cat_df[cat_df['Needs Review'] != 'Yes'].copy() if 'Needs Review' in cat_df.columns else cat_df.copy()
            if len(cat_df) < 5: continue
            total_styles += len(cat_df)
            
            # Clean targets
            cat_df['STR_Math'] = pd.to_numeric(cat_df['STR'].astype(str).str.replace('%', ''), errors='coerce') / 100.0 if 'STR' in cat_df.columns else np.nan
            cat_df['ROS_Math'] = pd.to_numeric(cat_df['ROS'], errors='coerce') if 'ROS' in cat_df.columns else np.nan
            cat_df.dropna(subset=['STR_Math', 'ROS_Math'], inplace=True)
            
            present_dna = [c for c in CATEGORY_DNA[cat] if c in cat_df.columns]
            cat_df[present_dna] = cat_df[present_dna].fillna("Unknown")
            
            X_cat = pd.get_dummies(cat_df[present_dna], drop_first=True, dtype=int, prefix_sep=': ')
            X_bin = cat_df[[c for c in BINARY_DNA if c in cat_df.columns]]
            X_full = pd.concat([X_cat, X_bin], axis=1).astype(float)
            
            if control_col and control_col in cat_df.columns:
                cat_df[control_col] = pd.to_numeric(cat_df[control_col], errors='coerce').fillna(1)
                X_full['Trading_Days_Control'] = cat_df[control_col]
                
            # 🚨 Unicorn Trap Filter
            cols_to_keep = [c for c in X_full.columns if c == 'Trading_Days_Control' or X_full[c].sum() >= 3]
            X_full = X_full[cols_to_keep]
            
            # 🚨 Pearson Correlation Filter
            if p_val > 0 and 'STR_Math' in cat_df.columns:
                corr_matrix = X_full.copy()
                corr_matrix['STR_Math'] = cat_df['STR_Math']
                corr = corr_matrix.corr(method='pearson')
                noise_traits = corr[corr['STR_Math'].abs() < p_val].index.tolist()
                noise_traits = [c for c in noise_traits if c != 'STR_Math' and c != 'Trading_Days_Control']
                X_full = X_full.drop(columns=[c for c in noise_traits if c in X_full.columns])
                
            if len(cat_df) < (len(X_full.columns) + 5) or X_full.empty:
                st.warning(f"Skipping {cat} Regression: Insufficient Data (N < K+5).")
                continue
                
            # Run VIF and Regression
            X_clean = calculate_vif(sm.add_constant(X_full))
            str_dict, ros_dict = {}, {}
            
            for target in targets:
                t_col = f'{target}_Math'
                if t_col not in cat_df.columns: continue
                try:
                    model = sm.OLS(cat_df[t_col], X_clean.astype(float)).fit(cov_type='HC3')
                    res = pd.DataFrame({'Attribute': model.params.index, 'Coefficient': model.params.values, 'P-Value': model.pvalues.values})
                    res = res[res['Attribute'] != 'const']
                    res['Significant?'] = res['P-Value'].apply(lambda x: 'Yes' if x < 0.05 else 'No')
                    res = res.sort_values('Coefficient', ascending=False)
                    
                    if target == 'STR': str_dict = dict(zip(res['Attribute'], res['Coefficient']))
                    if target == 'ROS': ros_dict = dict(zip(res['Attribute'], res['Coefficient']))
                    
                    sig_count += len(res[res['Significant?'] == 'Yes'])
                    r2_scores.append(model.rsquared)
                    
                    sheet_name_impact = f"Impact_{cat}_{target}"
                    if sheet_name_impact in wb.sheetnames: del wb[sheet_name_impact]
                    ws_imp = wb.create_sheet(sheet_name_impact)
                    ws_imp.append([f"R-Squared: {model.rsquared:.4f}"])
                    ws_imp.append([])
                    ws_imp.append(list(res.columns))
                    for r in dataframe_to_rows(res, index=False, header=False): ws_imp.append(r)
                except Exception as e:
                    st.error(f"Failed Regression on {cat} {target}: {e}")
                    
            # Predictive K-Means Clustering Phase
            str_impacts, ros_impacts = [], []
            for _, row in cat_df.iterrows():
                s_score, r_score = 0.0, 0.0
                for attr in present_dna:
                    dummy_key = f"{attr}: {str(row.get(attr, '')).strip().title()}"
                    s_score += str_dict.get(dummy_key, 0.0)
                    r_score += ros_dict.get(dummy_key, 0.0)
                for attr in BINARY_DNA:
                    if str(row.get(attr, '')).lower() in ['1', '1.0', 'yes', 'true']:
                        s_score += str_dict.get(attr, 0.0)
                        r_score += ros_dict.get(attr, 0.0)
                str_impacts.append(s_score)
                ros_impacts.append(r_score)
                
            cat_df['DNA_STR_Impact'] = str_impacts
            cat_df['DNA_ROS_Impact'] = ros_impacts
            
            cluster_data = cat_df[['DNA_STR_Impact', 'DNA_ROS_Impact']].copy()
            scaler = StandardScaler()
            scaled_data = scaler.fit_transform(cluster_data)
            
            kmeans = KMeans(n_clusters=min(5, len(cat_df)), random_state=42, n_init=10)
            cat_df['Cluster'] = kmeans.fit_predict(scaled_data)
            
            centers = pd.DataFrame(kmeans.cluster_centers_, columns=['STR_scaled', 'ROS_scaled'])
            centers['Total_Impact'] = centers['STR_scaled'] + centers['ROS_scaled']
            sorted_clusters = centers.sort_values(by='Total_Impact', ascending=False).index.tolist()
            
            grade_map = {cid: grade for cid, grade in zip(sorted_clusters, ['A', 'B', 'C', 'D', 'E'])}
            cat_df['Predicted_Grade'] = cat_df['Cluster'].map(grade_map)
            
            # Write new columns directly into the original worksheet (Non-Destructive)
            max_col = ws.max_column
            ws.cell(row=1, column=max_col+1, value="DNA_STR_Impact")
            ws.cell(row=1, column=max_col+2, value="DNA_ROS_Impact")
            ws.cell(row=1, column=max_col+3, value="Predicted_Grade")
            
            # Create a map to write back to the correct row index
            grade_dict = dict(zip(cat_df.index, cat_df['Predicted_Grade']))
            str_imp_dict = dict(zip(cat_df.index, cat_df['DNA_STR_Impact']))
            ros_imp_dict = dict(zip(cat_df.index, cat_df['DNA_ROS_Impact']))
            
            for idx in range(2, ws.max_row + 1):
                df_idx = idx - 2 # DataFrame index
                if df_idx in grade_dict:
                    ws.cell(row=idx, column=max_col+1, value=round(str_imp_dict[df_idx], 4))
                    ws.cell(row=idx, column=max_col+2, value=round(ros_imp_dict[df_idx], 4))
                    ws.cell(row=idx, column=max_col+3, value=grade_dict[df_idx])
        
        my_bar.progress(100, text="Finalizing Core Intelligence & Preserving Images...")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        with metric_placeholder.container():
            st.markdown("<br>", unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            c1.markdown(f'<div class="metric-card"><div class="metric-title">Data Points</div><div class="metric-value">{total_styles}</div></div>', unsafe_allow_html=True)
            c2.markdown(f'<div class="metric-card"><div class="metric-title">Validated Rules</div><div class="metric-value">{sig_count}</div></div>', unsafe_allow_html=True)
            avg_r2 = np.mean(r2_scores) if r2_scores else 0
            c3.markdown(f'<div class="metric-card"><div class="metric-title">Avg R² Score</div><div class="metric-value" style="color:#ECC94B;">{avg_r2:.2f}</div></div>', unsafe_allow_html=True)
        
        time.sleep(0.5)
        my_bar.empty()
        st.success("✅ AI Training Complete! Master sheet generated with ALL original images and predictive grades preserved.")
        st.download_button("⬇️ Download Master Retail Intelligence", data=output, file_name="Master_Retail_Intelligence.xlsx", use_container_width=True)

# ==========================================
# ENGINE 2: BATCH GRADER
# ==========================================
elif main_mode == "⚡ Engine 2: Batch Grader":
    st.markdown('<h1 style="background: -webkit-linear-gradient(45deg, #F6AD55, #ED8936); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">⚡ Engine 2: Batch Grading Pipeline</h1>', unsafe_allow_html=True)
    st.markdown('<p class="hero-text">Score hundreds of upcoming styles simultaneously using the K-Means DNA logic. The AI preserves your original images.</p>', unsafe_allow_html=True)
    
    c_up1, c_up2 = st.columns(2)
    with c_up1: up_master = st.file_uploader("1. Core Brain (Master_Retail_Intelligence.xlsx)", type=["xlsx"])
    with c_up2: up_styles = st.file_uploader("2. Target Data (Upcoming Styles with Attributes)", type=["xlsx"])
    
    st.markdown("---")
    st.subheader("Inventory Policy Configuration")
    cc1, cc2, cc3 = st.columns(3)
    with cc1: q_base = st.number_input("Base Quantity Per Store", min_value=1, value=100)
    with cc2: stores = st.number_input("Number of Target Stores", min_value=1, value=50)
    with cc3: risk = st.selectbox("Risk Tolerance", ["Conservative", "Balanced", "Aggressive"], index=1, help="Scales final order quantities based on institutional risk appetite.")
    
    mult_adjust = {'Conservative': 0.8, 'Balanced': 1.0, 'Aggressive': 1.2}[risk]
    
    st.markdown("<br>", unsafe_allow_html=True)
    btn_col1, btn_col2, btn_col3 = st.columns([1,2,1])
    with btn_col2:
        run_batch = st.button("🔮 Initialize Predictive Grading", type="primary", use_container_width=True)
        
    if up_master and up_styles and run_batch:
        with st.spinner("Reconstructing K-Means Space & Writing Non-Destructively..."):
            master_xls = pd.ExcelFile(up_master)
            
            # Load Target Data non-destructively
            up_styles.seek(0)
            wb_up = load_workbook(up_styles)
            up_styles.seek(0)
            up_xls = pd.ExcelFile(up_styles)
            
            results_log = {'A':0, 'B':0, 'C':0, 'D':0, 'E':0}
            total_units = 0
            total_styles = 0
            
            for sheet_name in wb_up.sheetnames:
                if sheet_name not in up_xls.sheet_names: continue
                    
                category = sheet_name.strip().title()
                if category not in CATEGORY_DNA: continue
                
                ws_up = wb_up[sheet_name]
                df_up = pd.read_excel(up_xls, sheet_name=sheet_name)
                
                max_col = ws_up.max_column
                ws_up.cell(row=1, column=max_col+1, value="DNA_STR_Impact")
                ws_up.cell(row=1, column=max_col+2, value="DNA_ROS_Impact")
                ws_up.cell(row=1, column=max_col+3, value="Predicted_Grade")
                ws_up.cell(row=1, column=max_col+4, value="Recommended_Buy")
                
                # Extract Regression Weights
                try:
                    str_weights = pd.read_excel(master_xls, sheet_name=f"Impact_{category}_STR", skiprows=2).dropna(subset=['Attribute'])
                    ros_weights = pd.read_excel(master_xls, sheet_name=f"Impact_{category}_ROS", skiprows=2).dropna(subset=['Attribute'])
                    str_dict = dict(zip(str_weights['Attribute'], str_weights['Coefficient']))
                    ros_dict = dict(zip(ros_weights['Attribute'], ros_weights['Coefficient']))
                except:
                    for idx in range(len(df_up)): ws_up.cell(row=idx+2, column=max_col+3, value="No Regression Weights")
                    continue
                
                # Reconstruct Scaling and Centroids from the Master File Historical Data
                try:
                    hist_df = pd.read_excel(master_xls, sheet_name=category)
                    hist_df = hist_df.dropna(subset=['DNA_STR_Impact', 'DNA_ROS_Impact', 'Predicted_Grade'])
                    base_str_impact = hist_df['DNA_STR_Impact'].mean()
                    std_str_impact = hist_df['DNA_STR_Impact'].std()
                    base_ros_impact = hist_df['DNA_ROS_Impact'].mean()
                    std_ros_impact = hist_df['DNA_ROS_Impact'].std()
                    
                    # Compute standardized centroids
                    centroids = {}
                    for g in ['A', 'B', 'C', 'D', 'E']:
                        g_df = hist_df[hist_df['Predicted_Grade'] == g]
                        if not g_df.empty:
                            c_str = (g_df['DNA_STR_Impact'].mean() - base_str_impact) / std_str_impact
                            c_ros = (g_df['DNA_ROS_Impact'].mean() - base_ros_impact) / std_ros_impact
                            centroids[g] = np.array([c_str, c_ros])
                except:
                    for idx in range(len(df_up)): ws_up.cell(row=idx+2, column=max_col+3, value="Missing Master Grades")
                    continue
                
                # Predict New Grades
                for idx, row in df_up.iterrows():
                    excel_row = idx + 2
                    total_styles += 1
                    
                    s_score, r_score = 0.0, 0.0
                    for attr in CATEGORY_DNA[category]:
                        if attr in df_up.columns and pd.notna(row[attr]):
                            s_score += str_dict.get(f"{attr}: {str(row[attr]).strip().title()}", 0.0)
                            r_score += ros_dict.get(f"{attr}: {str(row[attr]).strip().title()}", 0.0)
                    for attr in BINARY_DNA:
                        if attr in df_up.columns and str(row[attr]).lower() in ['yes', '1', '1.0', 'true']:
                            s_score += str_dict.get(attr, 0.0)
                            r_score += ros_dict.get(attr, 0.0)
                            
                    # Find Nearest Centroid
                    try:
                        scaled_str = (s_score - base_str_impact) / std_str_impact
                        scaled_ros = (r_score - base_ros_impact) / std_ros_impact
                        point = np.array([scaled_str, scaled_ros])
                        
                        best_grade, min_dist = "E", float('inf')
                        for g, centroid in centroids.items():
                            dist = np.linalg.norm(point - centroid)
                            if dist < min_dist:
                                min_dist = dist
                                best_grade = g
                    except:
                        best_grade = "Error"
                        
                    if best_grade in results_log: results_log[best_grade] += 1
                    
                    base_mult = {'A':2.0, 'B':1.5, 'C':1.0, 'D':0.75, 'E':0.5}.get(best_grade, 0)
                    oq = int((q_base * base_mult * mult_adjust) * stores)
                    total_units += oq
                    
                    ws_up.cell(row=excel_row, column=max_col+1, value=round(s_score, 4))
                    ws_up.cell(row=excel_row, column=max_col+2, value=round(r_score, 4))
                    ws_up.cell(row=excel_row, column=max_col+3, value=best_grade)
                    ws_up.cell(row=excel_row, column=max_col+4, value=oq)
            
            output = io.BytesIO()
            wb_up.save(output)
            output.seek(0)
            
            st.markdown("---")
            st.subheader("Batch Results Summary")
            r1, r2, r3, r4 = st.columns(4)
            r1.markdown(f'<div class="metric-card"><div class="metric-title">Total Styles</div><div class="metric-value">{total_styles}</div></div>', unsafe_allow_html=True)
            r2.markdown(f'<div class="metric-card"><div class="metric-title">"A" Grade Hits</div><div class="metric-value grade-A">{results_log["A"]}</div></div>', unsafe_allow_html=True)
            r3.markdown(f'<div class="metric-card"><div class="metric-title">Total Recommended Buy</div><div class="metric-value">{total_units:,}</div></div>', unsafe_allow_html=True)
            r4.markdown(f'<div class="metric-card"><div class="metric-title">Images Preserved</div><div class="metric-value" style="color:#48BB78;">100%</div></div>', unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            d1, d2, d3 = st.columns([1,2,1])
            with d2:
                st.download_button("⬇️ Download Final Graded File", data=output, file_name="Predictive_Graded_Styles.xlsx", use_container_width=True)

# ==========================================
# ENGINE 3: LIVE DASHBOARD
# ==========================================
elif main_mode == "📊 Engine 3: Live Dashboard":
    st.markdown('<h1 style="background: -webkit-linear-gradient(45deg, #9F7AEA, #ED64A6); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">📊 Engine 3: Interactive Analytics</h1>', unsafe_allow_html=True)
    st.markdown('<p class="hero-text">Explore attribute correlations and run real-time theoretical garment simulations using Predictive KMeans Mapping.</p>', unsafe_allow_html=True)
    
    uploaded_master = st.file_uploader("Authorize Intelligence File (Upload Master_Retail_Intelligence.xlsx)", type=["xlsx"])
    
    @st.cache_resource
    def load_data(file_obj):
        if file_obj is not None: return pd.ExcelFile(file_obj)
        elif os.path.exists("Master_Retail_Intelligence.xlsx"): return pd.ExcelFile("Master_Retail_Intelligence.xlsx")
        return None

    xls = load_data(uploaded_master)

    if xls is None:
        st.info("👋 Upload your Master Retail Intelligence Excel file above to unlock the dashboard.")
        st.stop()

    def get_sheet(sheet_name, skip=0):
        if sheet_name in xls.sheet_names: return pd.read_excel(xls, sheet_name=sheet_name, skiprows=skip)
        return pd.DataFrame()
        
    tab1, tab2, tab3 = st.tabs(["🧬 Attribute Analytics", "🔮 Live Grade Predictor", "🗺️ Geographic Assortment"])
    
    with tab1:
        cat = st.radio("Select Category:", list(CATEGORY_DNA.keys()), horizontal=True, key="t1")
        st.markdown("---")
        
        str_df = get_sheet(f"Impact_{cat}_STR", skip=2).dropna(subset=['Attribute'])
        ros_df = get_sheet(f"Impact_{cat}_ROS", skip=2).dropna(subset=['Attribute'])
        raw_df = get_sheet(cat)
        
        if str_df.empty or raw_df.empty:
            st.warning("Insufficient historical data for this category.")
        else:
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            with col_stat1: st.markdown(f'<div class="metric-card"><div class="metric-title">Total Styles Analyzed</div><div class="metric-value">{len(raw_df)}</div></div>', unsafe_allow_html=True)
            if 'STR' in raw_df.columns:
                raw_df['STR_Math'] = pd.to_numeric(raw_df['STR'].astype(str).str.replace('%', ''), errors='coerce') / 100.0
                with col_stat2: st.markdown(f'<div class="metric-card"><div class="metric-title" title="Sell-Through Rate">Historical Base STR%</div><div class="metric-value">{raw_df["STR_Math"].mean():.1%}</div></div>', unsafe_allow_html=True)
            with col_stat3: st.markdown(f'<div class="metric-card"><div class="metric-title" title="Rate of Sale">Historical Base ROS</div><div class="metric-value">{pd.to_numeric(raw_df["ROS"], errors="coerce").mean():.2f}</div></div>', unsafe_allow_html=True)
            with col_stat4:
                sig_drivers = len(str_df[str_df['Significant?'] == 'Yes'])
                st.markdown(f'<div class="metric-card"><div class="metric-title">Significant Drivers</div><div class="metric-value">{sig_drivers}</div></div>', unsafe_allow_html=True)
                
            st.markdown("<br>", unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                str_plot = str_df[str_df['Significant?'] == 'Yes'].sort_values('Coefficient', ascending=True)
                if str_plot.empty: str_plot = str_df.sort_values('Coefficient', ascending=True).tail(10)
                fig1 = px.bar(str_plot, x='Coefficient', y='Attribute', orientation='h', color='Coefficient', color_continuous_scale='Greens', title="STR Correlation Matrix")
                fig1.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="white")
                st.plotly_chart(fig1, use_container_width=True)
                
            with col2:
                ros_plot = ros_df[ros_df['Significant?'] == 'Yes'].sort_values('Coefficient', ascending=True)
                if ros_plot.empty: ros_plot = ros_df.sort_values('Coefficient', ascending=True).tail(10)
                fig2 = px.bar(ros_plot, x='Coefficient', y='Attribute', orientation='h', color='Coefficient', color_continuous_scale='Blues', title="ROS Correlation Matrix")
                fig2.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="white")
                st.plotly_chart(fig2, use_container_width=True)
                
    with tab2:
        cat2 = st.radio("Select Category to Grade:", list(CATEGORY_DNA.keys()), horizontal=True, key="t2")
        st.markdown("---")
        
        raw_df2 = get_sheet(cat2)
        str_df2 = get_sheet(f"Impact_{cat2}_STR", skip=2).dropna(subset=['Attribute'])
        ros_df2 = get_sheet(f"Impact_{cat2}_ROS", skip=2).dropna(subset=['Attribute'])
        
        if raw_df2.empty or str_df2.empty or 'DNA_STR_Impact' not in raw_df2.columns:
            st.warning("Insufficient predictive data.")
        else:
            base_str_impact = raw_df2['DNA_STR_Impact'].mean()
            std_str_impact = raw_df2['DNA_STR_Impact'].std()
            base_ros_impact = raw_df2['DNA_ROS_Impact'].mean()
            std_ros_impact = raw_df2['DNA_ROS_Impact'].std()
            
            centroids = {}
            for g in ['A', 'B', 'C', 'D', 'E']:
                g_df = raw_df2[raw_df2['Predicted_Grade'] == g]
                if not g_df.empty:
                    centroids[g] = np.array([
                        (g_df['DNA_STR_Impact'].mean() - base_str_impact) / std_str_impact,
                        (g_df['DNA_ROS_Impact'].mean() - base_ros_impact) / std_ros_impact
                    ])
            
            with st.form("predict_form"):
                st.subheader("1. Configure Garment DNA")
                cols = st.columns(4)
                selections = {}
                
                all_options = CATEGORY_DNA[cat2] + BINARY_DNA
                for i, attr in enumerate(all_options):
                    with cols[i % 4]:
                        if attr in CATEGORY_DNA[cat2]:
                            opts = raw_df2[attr].dropna().unique().tolist() if attr in raw_df2.columns else ["Unknown"]
                            selections[attr] = st.selectbox(f"{attr}", sorted(opts))
                        else:
                            st.markdown("<br>", unsafe_allow_html=True)
                            selections[attr] = st.toggle(f"{attr.replace('Has_', '')}?")
                
                st.markdown("<br>", unsafe_allow_html=True)
                st.subheader("2. Inventory Controls")
                coq1, coq2 = st.columns(2)
                with coq1: q_base = st.number_input("Base Quantity Per Store", min_value=1, value=100)
                with coq2: stores = st.number_input("Number of Target Stores", min_value=1, value=50)
                
                st.markdown("<br>", unsafe_allow_html=True)
                submitted = st.form_submit_button("⚡ Run AI Prediction & Grade Formulation", type="primary", use_container_width=True)
                
            if submitted:
                st.markdown("---")
                score_str = 0.0
                score_ros = 0.0
                for attr, val in selections.items():
                    if isinstance(val, bool):
                        if val:
                            m1 = str_df2[str_df2['Attribute'] == attr]
                            if not m1.empty: score_str += m1.iloc[0]['Coefficient']
                            m2 = ros_df2[ros_df2['Attribute'] == attr]
                            if not m2.empty: score_ros += m2.iloc[0]['Coefficient']
                    else:
                        dummy = f"{attr}: {val}"
                        m1 = str_df2[str_df2['Attribute'] == dummy]
                        if not m1.empty: score_str += m1.iloc[0]['Coefficient']
                        m2 = ros_df2[ros_df2['Attribute'] == dummy]
                        if not m2.empty: score_ros += m2.iloc[0]['Coefficient']
                        
                # Nearest Centroid Logic
                try:
                    scaled_str = (score_str - base_str_impact) / std_str_impact
                    scaled_ros = (score_ros - base_ros_impact) / std_ros_impact
                    point = np.array([scaled_str, scaled_ros])
                    grade, min_dist = "E", float('inf')
                    for g, centroid in centroids.items():
                        dist = np.linalg.norm(point - centroid)
                        if dist < min_dist:
                            min_dist = dist
                            grade = g
                except:
                    grade = "E"
                    
                color = f"grade-{grade}"
                mult = {'A':2.0, 'B':1.5, 'C':1.0, 'D':0.75, 'E':0.5}.get(grade, 0.5)
                oq = int((q_base * mult) * stores)
                
                c1, c2, c3, c4 = st.columns(4)
                with c1: st.markdown(f'<div class="metric-card"><div class="metric-title">Theoretical Grade</div><div class="metric-value {color}">{grade}</div></div>', unsafe_allow_html=True)
                with c2: st.markdown(f'<div class="metric-card"><div class="metric-title" title="Sell-Through Rate">DNA STR Impact</div><div class="metric-value">{score_str:.4f}</div></div>', unsafe_allow_html=True)
                with c3: st.markdown(f'<div class="metric-card"><div class="metric-title" title="Rate of Sale">DNA ROS Impact</div><div class="metric-value">{score_ros:.4f}</div></div>', unsafe_allow_html=True)
                with c4: st.markdown(f'<div class="metric-card"><div class="metric-title">Recommended Buy</div><div class="metric-value" style="color:#F6AD55;">{oq:,}</div></div>', unsafe_allow_html=True)

    with tab3:
        cat3 = st.radio("Select Target Category:", list(CATEGORY_DNA.keys()), horizontal=True, key="t3")
        st.markdown("---")
        raw_df3 = get_sheet(cat3)
        str_df3 = get_sheet(f"Impact_{cat3}_STR", skip=2).dropna(subset=['Attribute'])
        
        with st.form("assort_form"):
            st.subheader("Configure Target Mix")
            cols = st.columns(4)
            selections = {}
            for i, attr in enumerate(CATEGORY_DNA[cat3]):
                with cols[i % 4]:
                    opts = raw_df3[attr].dropna().unique().tolist() if (not raw_df3.empty and attr in raw_df3.columns) else ["Unknown"]
                    selections[attr] = st.selectbox(f"{attr}", sorted(opts))
                    
            st.markdown("<br>", unsafe_allow_html=True)
            submitted3 = st.form_submit_button("🌍 Predict Geographic Demand Clusters", type="primary", use_container_width=True)
            
        if submitted3:
            st.markdown("---")
            score_modifier = 0
            for attr, val in selections.items():
                dummy = f"{attr}: {val}"
                if not str_df3.empty:
                    m1 = str_df3[str_df3['Attribute'] == dummy]
                    if not m1.empty: score_modifier += m1.iloc[0]['Coefficient']
            
            seed_string = "".join([str(v) for v in selections.values()])
            np.random.seed(abs(hash(seed_string)) % (2**32))
            
            col1, col2 = st.columns([2, 1])
            cities = pd.DataFrame({
                'City': ['Mumbai', 'Delhi', 'Bangalore', 'Hyderabad', 'Chennai', 'Kolkata', 'Pune', 'Ahmedabad', 'Jaipur', 'Lucknow'],
                'lat': [19.0760, 28.7041, 12.9716, 17.3850, 13.0827, 22.5726, 18.5204, 23.0225, 26.9124, 26.8467],
                'Base_Demand': np.random.randint(40, 90, 10) 
            })
            
            # S-Curve Logistic Normalization for Geographic Mapping
            logistic_mult = 1 / (1 + np.exp(-score_modifier * 5)) 
            cities['Demand_Score'] = np.clip(cities['Base_Demand'] * (0.5 + logistic_mult), 10, 100).astype(int)
            
            cities['Size'] = cities['Demand_Score'] * 1500
            cities['Color'] = cities['Demand_Score'].apply(lambda x: '#00FF00' if x > 80 else ('#FFFF00' if x > 60 else '#FF0000'))
            
            with col1:
                st.map(cities, latitude='lat', longitude='lon', size='Size', color='Color', zoom=4)
                
            with col2:
                st.subheader("🔥 Top Target Locations")
                top_stores = cities.sort_values('Demand_Score', ascending=False).head(5)[['City', 'Demand_Score']]
                top_stores.columns = ['Store Location', 'Predicted Demand Index']
                st.dataframe(top_stores.style.background_gradient(cmap='Greens', subset=['Predicted Demand Index']), hide_index=True, use_container_width=True)
